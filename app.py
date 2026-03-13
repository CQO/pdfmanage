# -*- coding: utf-8 -*-
"""
功能1：表格识别V3 - 图片/PDF转Excel
功能2：图纸图号识别 - PDF图纸批量重命名
功能3：增值税发票识别 - 识别增值税发票关键信息
功能4：本地识别 - 调用本地OCR服务
"""
import os
import sys
import base64
import json
import re
import shutil
import time
import threading
import requests  # 新增导入
import openpyxl
from io import BytesIO
from pathlib import Path
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from typing import Union, List, Optional
import cv2
import numpy as np
from PIL import Image
import fitz  # PyMuPDF

# 腾讯云SDK
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.ocr.v20181119 import ocr_client, models

# ==================== 配置区域 ====================
# 从环境变量读取密钥（推荐）
TENCENT_SECRET_ID = "AKID62ub6KoNnDWkz50ymMq58mQxTp0161mO"
TENCENT_SECRET_KEY = "Zw9C5ttobWK0a5zztdDk6TjnnsxnRt8A"
DEFAULT_REGION = "ap-shanghai"  # 图纸识别推荐上海，表格识别推荐广州
TEMP_DIR = "temp_drawing"
OUTPUT_DIR = "./识别结果"

def setup_temp_dir():
    """创建临时目录"""
    if os.path.exists(TEMP_DIR):
        shutil.rmtree(TEMP_DIR)
    os.makedirs(TEMP_DIR, exist_ok=True)

def cleanup_temp():
    """清理临时目录"""
    if os.path.exists(TEMP_DIR):
        shutil.rmtree(TEMP_DIR)



# ==================== 增值税发票识别模块 ====================
class VatInvoiceRecognizer:
    """增值税发票识别类"""
    
    def __init__(self, secret_id=None, secret_key=None, region="ap-shanghai"):
        self.secret_id = secret_id or TENCENT_SECRET_ID
        self.secret_key = secret_key or TENCENT_SECRET_KEY
        self.region = region
        
    def recognize_invoice(self, image_path):
        """
        识别增值税发票
        返回: 发票信息的字典
        """
        try:
            # 初始化OCR客户端
            cred = credential.Credential(self.secret_id, self.secret_key)
            httpProfile = HttpProfile()
            httpProfile.endpoint = "ocr.tencentcloudapi.com"
            clientProfile = ClientProfile()
            clientProfile.httpProfile = httpProfile
            client = ocr_client.OcrClient(cred, self.region, clientProfile)
            
            # 读取图片
            with open(image_path, 'rb') as f:
                base64_data = base64.b64encode(f.read())
                img_base64 = base64_data.decode()
            
            # 调用增值税发票识别
            req = models.VatInvoiceOCRRequest()
            params = '{"ImageBase64":"%s"}' % img_base64
            req.from_json_string(params)
            
            resp = client.VatInvoiceOCR(req)
            resp = json.loads(resp.to_json_string())
            
            # 解析识别结果
            invoice_info = {}
            for item in resp.get('VatInvoiceInfos', []):
                name = item.get('Name', '')
                value = item.get('Value', '')
                invoice_info[name] = value
            
            return invoice_info
            
        except Exception as e:
            raise Exception(f"发票识别失败: {str(e)}")
    
    def format_invoice_info(self, invoice_info):
        """格式化发票信息为可读文本"""
        lines = []
        lines.append("=" * 60)
        lines.append("增值税发票识别结果")
        lines.append("=" * 60)
        
        # 定义需要显示的字段及其中文名称
        fields = {
            'Name': '名称',
            'Type': '类型',
            'Code': '发票代码',
            'Number': '发票号码',
            'Date': '开票日期',
            'Total': '金额',
            'TotalCn': '金额(大写)',
            'Tax': '税额',
            'Amount': '价税合计',
            'AmountCn': '价税合计(大写)',
            'CheckCode': '校验码',
            'SellerName': '销售方名称',
            'SellerTaxID': '销售方税号',
            'BuyerName': '购买方名称',
            'BuyerTaxID': '购买方税号',
            'Remark': '备注'
        }
        
        for key, ch_name in fields.items():
            if key in invoice_info and invoice_info[key]:
                lines.append(f"{ch_name}: {invoice_info[key]}")
        
        lines.append("=" * 60)
        return "\n".join(lines)
    
    def save_as_json(self, invoice_info, output_path):
        """保存为JSON文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(invoice_info, f, ensure_ascii=False, indent=2)
        return output_path
    
    def save_as_txt(self, invoice_info, output_path):
        """保存为TXT文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(self.format_invoice_info(invoice_info))
        return output_path
    
    def process_invoice(self, image_path, output_format='both', log_callback=None):
        """
        处理单个发票文件
        output_format: 'json', 'txt', 或 'both'
        """
        def log(msg):
            if log_callback:
                log_callback(msg)
            else:
                print(msg)
        
        try:
            log(f"📄 处理发票: {os.path.basename(image_path)}")
            
            # 识别发票
            invoice_info = self.recognize_invoice(image_path)
            
            if not invoice_info:
                log("❌ 未识别到发票信息")
                return None
            
            # 获取发票号码用于文件名
            invoice_number = invoice_info.get('Number', '未知发票')
            base_name = Path(image_path).stem
            
            saved_files = []
            
            # 保存JSON
            if output_format in ['json', 'both']:
                json_path = os.path.join(OUTPUT_DIR, f"{base_name}_{invoice_number}.json")
                counter = 1
                original_json = json_path
                while os.path.exists(json_path):
                    name_part = Path(original_json).stem
                    if name_part.endswith(f"_{counter-1}"):
                        name_part = name_part[:-3]
                    json_path = os.path.join(OUTPUT_DIR, f"{name_part}_{counter}.json")
                    counter += 1
                
                self.save_as_json(invoice_info, json_path)
                saved_files.append(json_path)
                log(f"💾 已保存JSON: {os.path.basename(json_path)}")
            
            # 保存TXT
            if output_format in ['txt', 'both']:
                txt_path = os.path.join(OUTPUT_DIR, f"{base_name}_{invoice_number}.txt")
                counter = 1
                original_txt = txt_path
                while os.path.exists(txt_path):
                    name_part = Path(original_txt).stem
                    if name_part.endswith(f"_{counter-1}"):
                        name_part = name_part[:-3]
                    txt_path = os.path.join(OUTPUT_DIR, f"{name_part}_{counter}.txt")
                    counter += 1
                
                self.save_as_txt(invoice_info, txt_path)
                saved_files.append(txt_path)
                log(f"💾 已保存TXT: {os.path.basename(txt_path)}")
            
            # 显示识别结果摘要
            log(f"\n📊 识别结果摘要:")
            if 'SellerName' in invoice_info:
                log(f"   销售方: {invoice_info['SellerName']}")
            if 'BuyerName' in invoice_info:
                log(f"   购买方: {invoice_info['BuyerName']}")
            if 'Amount' in invoice_info:
                log(f"   价税合计: {invoice_info['Amount']}")
            if 'Date' in invoice_info:
                log(f"   开票日期: {invoice_info['Date']}")
            
            return saved_files
            
        except Exception as e:
            log(f"❌ 处理失败: {str(e)}")
            return None



# ==================== 表格识别模块 ====================
class TableOCRRecognizer:
    """表格识别V3封装类"""
    
    def __init__(self, secret_id=None, secret_key=None, region="ap-guangzhou"):
        self.secret_id = secret_id or TENCENT_SECRET_ID
        self.secret_key = secret_key or TENCENT_SECRET_KEY
        self.region = region
    
    def recognize_from_image(self, image_input):
        """表格识别V3核心方法"""
        # 实例化认证对象
        cred = credential.Credential(self.secret_id, self.secret_key)
        
        # HTTP配置
        http_profile = HttpProfile()
        http_profile.endpoint = "ocr.tencentcloudapi.com"
        http_profile.reqTimeout = 60
        
        # 客户端配置
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        client_profile.signMethod = "TC3-HMAC-SHA256"
        
        # 初始化客户端
        client = ocr_client.OcrClient(cred, self.region, client_profile)
        
        # 处理图片输入
        if isinstance(image_input, str):
            with open(image_input, 'rb') as f:
                img_data = f.read()
        else:
            img_data = image_input
        
        # 构造请求
        req = models.RecognizeTableAccurateOCRRequest()
        req.ImageBase64 = base64.b64encode(img_data).decode('utf-8')
        
        # PDF处理
        if isinstance(image_input, str) and image_input.lower().endswith('.pdf'):
            req.PdfPageNumber = 1
        
        # 发起请求
        resp = client.RecognizeTableAccurateOCR(req)
        excel_data = base64.b64decode(resp.Data)
        return excel_data


    def replace_in_excel_file(self, excel_data, pattern_replacements):
        """
        使用openpyxl处理Excel文件，安全地替换单元格内容
        pattern_replacements: 列表，每个元素为 (正则表达式, 替换字符串或函数)
        """
        try:
            # 将二进制数据加载为Excel工作簿
            excel_bytes = BytesIO(excel_data)
            wb = openpyxl.load_workbook(excel_bytes)
            
            # 遍历所有工作表
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # 遍历所有单元格
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            original = cell.value
                            
                            # 应用所有正则规则
                            for pattern, repl in pattern_replacements:
                                cell.value = re.sub(pattern, repl, cell.value)
                            
                            # 如果发生变化，打印日志
                            if cell.value != original:
                                print(f"替换: '{original}' → '{cell.value}'")
            
            # 保存到新的BytesIO对象
            output_bytes = BytesIO()
            wb.save(output_bytes)
            output_bytes.seek(0)
            return output_bytes.read()
            
        except Exception as e:
            print(f"Excel处理失败: {e}")
            return excel_data

    def save_as_excel(self, image_input, output_path=None):
        """识别并保存为Excel文件"""
        excel_data = self.recognize_from_image(image_input)
        
        # ===== 使用正则表达式替换"中+任意数字"为"Φ+相同数字" =====
        pattern_replacements = [
            (r'中(\d)', r'Φ\1'),  # 中6 → Φ6, 中123 → Φ123
            # 可以添加更多正则规则
            # (r'直径(\d+)', r'Φ\1'),  # 直径6 → Φ6
        ]
        
        excel_data = self.replace_in_excel_file(excel_data, pattern_replacements)
        
        # 后续代码保持不变...
        if output_path is None:
            if isinstance(image_input, str):
                base_name = Path(image_input).stem
                output_path = f"{base_name}_识别结果.xlsx"
            else:
                output_path = "表格识别结果.xlsx"
        elif not output_path.endswith(('.xlsx', '.xls')):
            output_path += '.xlsx'
        
        # 处理重名
        counter = 1
        original_path = output_path
        while os.path.exists(output_path):
            name_part = Path(original_path).stem
            ext = Path(original_path).suffix
            if name_part.endswith(f"_{counter-1}"):
                name_part = name_part[:-3]
            output_path = f"{name_part}_{counter}{ext}"
            counter += 1
        
        with open(output_path, 'wb') as f:
            f.write(excel_data)
        
        return output_path


# ==================== 图纸图号识别模块 ====================
class DrawingNumberRecognizer:
    """图纸图号识别类（基于原代码优化）"""
    
    def __init__(self, secret_id=None, secret_key=None, region="ap-shanghai"):
        self.secret_id = secret_id or TENCENT_SECRET_ID
        self.secret_key = secret_key or TENCENT_SECRET_KEY
        self.region = region
        
    def cv_imread(self, file_path):
        """解决imread不能读取中文路径的问题"""
        cv_img = cv2.imdecode(np.fromfile(file_path, dtype=np.uint8), -1)
        return cv_img
    
    def change_image(self, img, angle):
        """旋转图像"""
        if angle == 90:
            img = cv2.transpose(img)
            img = cv2.flip(img, flipCode=1)
            return img
        # 任意角度旋转
        cx, cy = img.shape[1] / 2, img.shape[0] / 2
        M = cv2.getRotationMatrix2D((cx, cy), angle, 1)
        rotated_img = cv2.warpAffine(img, M, (img.shape[1], img.shape[0]))
        return rotated_img
    
    def pdf_to_image(self, pdf_path, zoom=3):
        """PDF转PNG图片"""
        pdf = fitz.open(pdf_path)
        images = []
        for pg in range(len(pdf)):
            page = pdf[pg]
            trans = fitz.Matrix(zoom, zoom)
            pm = page.get_pixmap(matrix=trans, alpha=False)
            img_path = os.path.join(TEMP_DIR, f"{Path(pdf_path).stem}_p{pg+1}.png")
            pm.save(img_path)
            images.append(img_path)
        pdf.close()
        return images
    
    def recognize_drawing_number(self, image_path):
        """
        识别图纸中的图号
        返回: [图号, 版本号, 部件标识] 如 ["DRM-2023-001", "-1", "-A"]
        """
        try:
            # 初始化OCR客户端
            cred = credential.Credential(self.secret_id, self.secret_key)
            httpProfile = HttpProfile()
            httpProfile.endpoint = "ocr.tencentcloudapi.com"
            clientProfile = ClientProfile("TC3-HMAC-SHA256")
            clientProfile.httpProfile = httpProfile
            client = ocr_client.OcrClient(cred, self.region, clientProfile)
            
            # 读取图片
            with open(image_path, 'rb') as f:
                base64_data = base64.b64encode(f.read())
                s = base64_data.decode()
            
            # 调用通用OCR
            req = models.GeneralBasicOCRRequest()
            params = '{"ImageBase64":"%s"}' % s
            req.from_json_string(params)
            
            resp = client.GeneralBasicOCR(req)
            resp = json.loads(resp.to_json_string())
            
            # 解析识别结果
            resp_list = resp.get('TextDetections', [])
            
            # 提取图号特征
            str0, str1, str2 = '', '-A', ''
            
            # 先识别版本标识
            for resp in resp_list:
                text = resp.get('DetectedText', '')
                match = re.findall(r'[1-9]/[1-9]', text)
                if len(match) >= 1:
                    str2 = match[0].replace('/', '-')
                    str2 = '-' + str2.split('-')[0]
                    if str2 == '-1' and str2.split('-')[1][0] != '2':
                        str2 = ''
                if 'A' in text:
                    str1 = '-A'
                if 'B' in text:
                    str1 = '-B'
            
            # 识别图号主体
            for resp in resp_list:
                result = resp.get('DetectedText', '')
                
                # 图号特征匹配
                if (('RM' in result and '-' in result) or 
                    ('ME' in result and '-' in result) or 
                    ('TF' in result and '.' in result) or 
                    (result.count('.') == 4)):
                    
                    result = result.replace(')', '1')
                    result = result.replace('图', '')
                    result = result.replace('号', '')
                    result = result.replace('专', '')
                    result = result.replace('+', '')
                    result = result.replace(' ', '|')
                    
                    result_parts = result.split('|')
                    for item in result_parts:
                        if 'ME' in item and '-' in item:
                            str0 = item
                        if 'RM' in item and '-' in item:
                            str0 = item
                        if 'TF' in item and '.' in item:
                            str0 = item
                        if item.count('.') == 4:
                            str0 = item
                    
                    if 'TF' in str0 and str0.find('TF') != 0:
                        str0 = str0[str0.find('TF'):]
                    if 'R' in str0 and 'DR' not in str0:
                        str0 = str0.replace('R', 'DR')
            
            # 清理结果
            if 'A' in str0:
                str0 = str0.split('A')[0]
            if 'B' in str0:
                str0 = str0.split('B')[0]
            
            str0 = str0.replace('/', '').replace('.', ' ').replace(':', ' ')
            str1 = str1.replace('.', ' ').replace(':', ' ')
            str2 = str2.replace('.', ' ').replace(':', ' ')
            
            return [str0.strip(), str2.strip(), str1.strip()]
            
        except Exception as e:
            print(f"识别失败: {str(e)}")
            return [None, None, None]
    
    def process_pdf_drawing(self, pdf_path, log_callback=None):
        """
        处理单个PDF图纸文件
        """
        def log(msg):
            if log_callback:
                log_callback(msg)
            else:
                print(msg)
        
        try:
            log(f"📄 处理图纸: {os.path.basename(pdf_path)}")
            
            # PDF转图片
            img_paths = self.pdf_to_image(pdf_path, zoom=3)
            if not img_paths:
                log("❌ PDF转图片失败")
                return None
            
            # 处理第一页（通常图号在第一页）
            img_path = img_paths[0]
            
            # 读取图片并调整方向
            img_big = self.cv_imread(img_path)
            if img_big is None:
                log("❌ 无法读取图片")
                return None
            
            height, width = img_big.shape[:2]
            
            # 如果宽度小于高度，先旋转90度
            if width < height:
                img_big = self.change_image(img_big, 90)
                height, width = img_big.shape[:2]
            
            # 裁剪右下角区域（图号通常在这里）
            crop_x = int(width * 0.45)
            crop_y = int(height * 0.7)
            img_crop = img_big[crop_y:height, crop_x:width]
            
            # 保存裁剪图片
            crop_path = os.path.join(TEMP_DIR, f"crop_{Path(img_path).name}")
            cv2.imencode('.png', img_crop)[1].tofile(crop_path)
            
            # 尝试多次旋转识别
            angles_to_try = [0, 180, 90, 270]
            best_result = [None, None, None]
            
            for angle in angles_to_try:
                if angle > 0:
                    rotated_img = self.change_image(img_big.copy(), angle)
                    height, width = rotated_img.shape[:2]
                    crop_x = int(width * 0.45)
                    crop_y = int(height * 0.7)
                    img_crop = rotated_img[crop_y:height, crop_x:width]
                    cv2.imencode('.png', img_crop)[1].tofile(crop_path)
                
                result = self.recognize_drawing_number(crop_path)
                if result[0] and result[0] not in ['', None]:
                    best_result = result
                    log(f"✅ 识别到图号: {result[0]}{result[1]}{result[2]}")
                    break
            
            if best_result[0]:
                # 生成新文件名
                new_filename = f"{best_result[0]}{best_result[1]}{best_result[2]}.pdf"
                new_path = os.path.join(OUTPUT_DIR, new_filename)
                
                # 处理重名
                counter = 1
                while os.path.exists(new_path):
                    name_part = f"{best_result[0]}{best_result[1]}{best_result[2]}"
                    new_filename = f"{name_part}_{counter}.pdf"
                    new_path = os.path.join(OUTPUT_DIR, new_filename)
                    counter += 1
                
                # 复制并重命名文件
                shutil.copy2(pdf_path, new_path)
                log(f"💾 已保存: {new_filename}")
                return new_path
            else:
                log("❌ 未识别到图号")
                return None
                
        except Exception as e:
            log(f"❌ 处理失败: {str(e)}")
            return None



# ==================== 通用文字识别模块 ====================
class GeneralOCRRecognizer:
    """通用文字识别类"""
    
    def __init__(self, secret_id=None, secret_key=None, region="ap-shanghai"):
        self.secret_id = secret_id or TENCENT_SECRET_ID
        self.secret_key = secret_key or TENCENT_SECRET_KEY
        self.region = region
    
    def recognize_text(self, image_path, language="auto", scene="general"):
        """
        通用文字识别
        language: 语言类型 (auto, zh, en, japan, korean 等)
        scene: 场景 (general, print, handwriting)
        返回: 识别的文本列表
        """
        try:
            # 初始化OCR客户端
            cred = credential.Credential(self.secret_id, self.secret_key)
            httpProfile = HttpProfile()
            httpProfile.endpoint = "ocr.tencentcloudapi.com"
            clientProfile = ClientProfile()
            clientProfile.httpProfile = httpProfile
            client = ocr_client.OcrClient(cred, self.region, clientProfile)
            
            # 读取图片
            with open(image_path, 'rb') as f:
                base64_data = base64.b64encode(f.read())
                img_base64 = base64_data.decode()
            
            # 根据场景选择不同的接口
            if scene == "handwriting":
                # 手写体识别
                req = models.GeneralHandwritingOCRRequest()
                params = '{"ImageBase64":"%s"}' % img_base64
                req.from_json_string(params)
                resp = client.GeneralHandwritingOCR(req)
            else:
                # 通用印刷体识别（支持多语言）
                req = models.GeneralBasicOCRRequest()
                params = {
                    "ImageBase64": img_base64,
                    "LanguageType": language
                }
                req.from_json_string(json.dumps(params))
                resp = client.GeneralBasicOCR(req)
            
            resp = json.loads(resp.to_json_string())
            
            # 解析识别结果
            text_items = []
            for item in resp.get('TextDetections', []):
                text = item.get('DetectedText', '')
                if text:
                    text_items.append(text)
            
            return text_items
            
        except Exception as e:
            raise Exception(f"文字识别失败: {str(e)}")
    
    def format_text_result(self, text_items):
        """格式化识别结果为文本"""
        if not text_items:
            return "未识别到文字"
        
        return "\n".join(text_items)
    
    def save_as_json(self, text_items, output_path):
        """保存为JSON文件"""
        result = {
            "text_count": len(text_items),
            "texts": text_items,
            "full_text": "\n".join(text_items)
        }
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        return output_path
    
    def save_as_txt(self, text_items, output_path):
        """保存为TXT文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(self.format_text_result(text_items))
        return output_path
    
    def process_image(self, image_path, output_dir, output_format='both', language='auto', scene='general', log_callback=None):
        """
        处理单个图片文件
        """
        def log(msg):
            if log_callback:
                log_callback(msg)
            else:
                print(msg)
        
        try:
            log(f"📄 处理图片: {os.path.basename(image_path)}")
            
            # 识别文字
            text_items = self.recognize_text(image_path, language, scene)
            
            if not text_items:
                log("❌ 未识别到文字")
                return None
            
            log(f"✅ 识别到 {len(text_items)} 段文字")
            
            # 显示前几行预览
            preview_lines = text_items[:5]
            log(f"📝 预览:")
            for line in preview_lines:
                if len(line) > 50:
                    line = line[:50] + "..."
                log(f"   {line}")
            if len(text_items) > 5:
                log(f"   ... 等 {len(text_items)} 段文字")
            
            # 生成输出文件名
            base_name = Path(image_path).stem
            saved_files = []
            
            # 保存JSON
            if output_format in ['json', 'both']:
                json_path = os.path.join(output_dir, f"{base_name}_文字识别.json")
                counter = 1
                original_json = json_path
                while os.path.exists(json_path):
                    name_part = Path(original_json).stem
                    if name_part.endswith(f"_{counter-1}"):
                        name_part = name_part[:-3]
                    json_path = os.path.join(output_dir, f"{name_part}_{counter}.json")
                    counter += 1
                
                self.save_as_json(text_items, json_path)
                saved_files.append(json_path)
                log(f"💾 已保存JSON: {os.path.basename(json_path)}")
            
            # 保存TXT
            if output_format in ['txt', 'both']:
                txt_path = os.path.join(output_dir, f"{base_name}_文字识别.txt")
                counter = 1
                original_txt = txt_path
                while os.path.exists(txt_path):
                    name_part = Path(original_txt).stem
                    if name_part.endswith(f"_{counter-1}"):
                        name_part = name_part[:-3]
                    txt_path = os.path.join(output_dir, f"{name_part}_{counter}.txt")
                    counter += 1
                
                self.save_as_txt(text_items, txt_path)
                saved_files.append(txt_path)
                log(f"💾 已保存TXT: {os.path.basename(txt_path)}")
            
            return saved_files
            
        except Exception as e:
            log(f"❌ 处理失败: {str(e)}")
            return None


# ==================== 本地识别模块 ====================
class LocalOCRRecognizer:
    """本地OCR识别类（调用本地服务）"""
    
    def __init__(self, api_url="http://127.0.0.1:10001/ocr"):
        self.api_url = api_url
    
    def recognize_image(self, image_path):
        """
        调用本地OCR服务识别图片
        返回: 识别的文本列表
        """
        try:
            # 读取图片文件
            with open(image_path, 'rb') as f:
                files = {'image': (os.path.basename(image_path), f, 'image/jpeg')}
                
                # 发送POST请求
                response = requests.post(self.api_url, files=files, timeout=30)
                
                # 检查响应状态
                if response.status_code != 200:
                    raise Exception(f"HTTP错误: {response.status_code}")
                
                # 解析JSON响应
                result = response.json()
                
                # 检查是否成功
                if not result.get('success', False):
                    raise Exception("服务返回失败状态")
                
                # 提取识别到的文字
                texts = []
                for item in result.get('texts', []):
                    rec_texts = item.get('rec_texts', [])
                    texts.extend(rec_texts)
                
                return texts
                
        except requests.exceptions.ConnectionError:
            raise Exception("无法连接到本地服务，请确认服务是否启动")
        except requests.exceptions.Timeout:
            raise Exception("请求超时")
        except json.JSONDecodeError:
            raise Exception("服务返回非JSON格式数据")
        except Exception as e:
            raise Exception(f"识别失败: {str(e)}")
    
    def format_text_result(self, text_items):
        """格式化识别结果为文本"""
        if not text_items:
            return "未识别到文字"
        
        return "\n".join(text_items)
    
    def save_as_txt(self, text_items, output_path):
        """保存为TXT文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(self.format_text_result(text_items))
        return output_path
    
    def process_image(self, image_path, output_dir, log_callback=None):
        """
        处理单个图片文件
        """
        def log(msg):
            if log_callback:
                log_callback(msg)
            else:
                print(msg)
        
        try:
            log(f"📄 处理图片: {os.path.basename(image_path)}")
            log(f"🔗 请求地址: {self.api_url}")
            
            # 识别文字
            text_items = self.recognize_image(image_path)
            
            if not text_items:
                log("❌ 未识别到文字")
                return None
            
            log(f"✅ 识别到 {len(text_items)} 段文字")
            
            # 显示前几行预览
            preview_lines = text_items[:5]
            log(f"📝 预览:")
            for line in preview_lines:
                if len(line) > 50:
                    line = line[:50] + "..."
                log(f"   {line}")
            if len(text_items) > 5:
                log(f"   ... 等 {len(text_items)} 段文字")
            
            # 生成输出文件名
            base_name = Path(image_path).stem
            txt_path = os.path.join(output_dir, f"{base_name}_本地识别.txt")
            
            # 处理重名
            counter = 1
            original_path = txt_path
            while os.path.exists(txt_path):
                name_part = Path(original_path).stem
                if name_part.endswith(f"_{counter-1}"):
                    name_part = name_part[:-3]
                txt_path = os.path.join(output_dir, f"{name_part}_{counter}.txt")
                counter += 1
            
            # 保存TXT
            self.save_as_txt(text_items, txt_path)
            log(f"💾 已保存TXT: {os.path.basename(txt_path)}")
            
            return txt_path
            
        except Exception as e:
            log(f"❌ 处理失败: {str(e)}")
            return None


# ==================== 主GUI应用 ====================
class OCRTabbedApp:
    """五选项卡OCR综合工具"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("CISDI文字识别工具")
        self.root.geometry("900x620")
        self.root.minsize(800, 620)

        # 共享变量
        self.secret_id = StringVar(value=TENCENT_SECRET_ID)
        self.secret_key = StringVar(value=TENCENT_SECRET_KEY)
        self.table_region = StringVar(value="ap-guangzhou")
        self.drawing_region = StringVar(value="ap-shanghai")
        self.invoice_region = StringVar(value="ap-shanghai")
        self.general_region = StringVar(value="ap-shanghai")
        
        # 本地识别服务地址
        self.local_api_url = StringVar(value="http://127.0.0.1:10001/ocr")
        
        # 发票输出格式
        self.invoice_format = StringVar(value="both")
        
        # 文字识别输出格式
        self.general_format = StringVar(value="txt")

        # 设置窗口图标
        try:
            self.root.iconbitmap('app_icon.ico')
        except Exception as e:
            print(f"图标加载失败: {e}")
        
        # 设置UI
        self.setup_ui()
    
    def setup_ui(self):
        """初始化用户界面"""
        # 创建菜单栏
        self.setup_menu()
        
        # 主框架
        main_frame = ttk.Frame(self.root, padding="5")
        main_frame.grid(row=0, column=0, sticky=(N, W, E, S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # ========== 选项卡 ==========
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=0, column=0, sticky=(N, S, E, W), pady=(0, 5))
        main_frame.rowconfigure(1, weight=1)
        
        # 创建五个选项卡
        self.setup_general_tab()    # 文字识别选项卡
        self.setup_drawing_tab()    # 图纸识别选项卡
        self.setup_table_tab()      # 表格识别选项卡
        self.setup_invoice_tab()    # 发票识别选项卡
        self.setup_local_tab()       # 本地识别选项卡（新增）

        # ========== 进度条和开始按钮区域 ==========
        control_frame = ttk.Frame(main_frame)
        control_frame.grid(row=1, column=0, pady=(0, 3), sticky=(W, E))
        
        # 开始按钮放在左边
        self.start_btn = ttk.Button(
            control_frame,
            text="开始识别",
            command=self.start_recognition,
            width=12
        )
        self.start_btn.grid(row=0, column=0, padx=(0, 10))
        
        # 进度条
        self.common_progress = ttk.Progressbar(control_frame, mode='determinate', length=250)
        self.common_progress.grid(row=0, column=1, padx=(0, 0))
        
        # 让进度条右侧的空间可扩展
        control_frame.columnconfigure(1, weight=1)
        
        # ========== 底部：公用日志 ==========
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="5")
        log_frame.grid(row=2, column=0, sticky=(W, E, S), pady=(0, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # 日志文本框
        self.common_log = Text(log_frame, height=15, wrap=WORD)
        self.common_log.grid(row=0, column=0, sticky=(N, S, E, W))
        
        log_scrollbar = ttk.Scrollbar(log_frame, orient=VERTICAL, command=self.common_log.yview)
        log_scrollbar.grid(row=0, column=1, sticky=(N, S))
        self.common_log.configure(yscrollcommand=log_scrollbar.set)
        
        # ========== 清除日志按钮放在日志框下面靠右侧 ==========
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, pady=(3, 0), sticky=(E))
        
        ttk.Button(
            button_frame,
            text="清除日志",
            command=self.clear_log,
            width=10
        ).grid(row=0, column=0)
    
    def setup_menu(self):
        """设置菜单栏"""
        menubar = Menu(self.root)
        self.root.config(menu=menubar)
        
        # 文件菜单
        file_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="文件", menu=file_menu)
        file_menu.add_command(label="退出", command=self.root.quit)
        
        # 设置菜单
        settings_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="设置", menu=settings_menu)
        
        # 导出设置子菜单
        export_menu = Menu(settings_menu, tearoff=0)
        settings_menu.add_cascade(label="导出设置", menu=export_menu)
        
        # 当前输出目录显示
        self.current_output_var = StringVar(value="未设置（使用默认目录）")
        export_menu.add_command(
            label="当前输出目录", 
            command=self.show_current_output,
            state=DISABLED
        )
        export_menu.add_separator()
        
        # 选择输出目录
        export_menu.add_command(
            label="选择输出目录", 
            command=self.select_common_output
        )
        
        # 清除输出目录
        export_menu.add_command(
            label="清除输出目录", 
            command=self.clear_common_output
        )


    # ========== 文字识别方法 ==========
    def select_general_files(self):
        """选择文字识别图片文件"""
        files = filedialog.askopenfilenames(
            title="选择图片文件",
            filetypes=[
                ("图像文件", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"),
                ("所有文件", "*.*")
            ]
        )
        
        if files:
            for file in files:
                if file not in self.general_files:
                    self.general_files.append(file)
                    self.general_listbox.insert(END, os.path.basename(file))
            
            self.general_file_label.config(text=f"已选择 {len(self.general_files)} 个文件")
            self.log(f"📎 文字识别文件: 已添加 {len(files)} 个文件，当前共 {len(self.general_files)} 个文件")

    def clear_general_files(self):
        """清空文字识别文件列表"""
        self.general_files.clear()
        self.general_listbox.delete(0, END)
        self.general_file_label.config(text="未选择文件")
        self.log("🗑️ 已清空文字识别文件列表")

    def setup_general_tab(self):
        """文字识别选项卡（通用OCR）"""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="文字识别")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)
        
        # ===== 文件选择 =====
        file_frame = ttk.LabelFrame(tab, text="图片文件", padding="5")
        file_frame.grid(row=0, column=0, sticky=(W, E), pady=(0, 5))
        file_frame.columnconfigure(1, weight=1)
        
        
        # 文字识别文件变量
        self.general_files = []
        
        ttk.Button(
            file_frame,
            text="📁 选择图片",
            command=self.select_general_files,
            width=12
        ).grid(row=0, column=0, padx=(0, 5))
        
        self.general_file_label = ttk.Label(file_frame, text="未选择文件")
        self.general_file_label.grid(row=0, column=1, sticky=W)
        
        ttk.Button(
            file_frame,
            text="清空",
            command=self.clear_general_files,
            width=6
        ).grid(row=0, column=2, padx=(5, 0))
        
        # 文件列表
        self.general_listbox = Listbox(
            file_frame,
            height=3,
            selectmode=EXTENDED,
            activestyle='none'
        )
        self.general_listbox.grid(row=1, column=0, columnspan=3, sticky=(W, E), pady=(5, 0))
        
        # ===== 输出格式设置 =====
        format_frame = ttk.LabelFrame(tab, text="输出格式设置", padding="5")
        format_frame.grid(row=1, column=0, sticky=(W, E), pady=(0, 5))
        
        ttk.Label(format_frame, text="保存格式:").grid(row=0, column=0, padx=(0, 5))
        
        ttk.Radiobutton(
            format_frame,
            text="TXT",
            variable=self.general_format,
            value="txt"
        ).grid(row=0, column=1, padx=2)
        
        ttk.Radiobutton(
            format_frame,
            text="JSON",
            variable=self.general_format,
            value="json"
        ).grid(row=0, column=2, padx=2)
        
        ttk.Radiobutton(
            format_frame,
            text="两者都保存",
            variable=self.general_format,
            value="both"
        ).grid(row=0, column=3, padx=2)
        
        # ===== 识别参数设置 =====
        param_frame = ttk.LabelFrame(tab, text="识别参数", padding="5")
        param_frame.grid(row=2, column=0, sticky=(W, E), pady=(0, 5))
        
        # 语言类型
        ttk.Label(param_frame, text="语言类型:").grid(row=0, column=0, padx=(0, 5))
        self.general_language = StringVar(value="auto")
        lang_combo = ttk.Combobox(
            param_frame,
            textvariable=self.general_language,
            values=["auto", "zh", "zh_auto", "en", "jap", "kor", "fre", "ger", "spa", "por", "vie", "may", "rus", "ita", "hol", "swe", "fin", "dan", "nor", "hun", "tha"],
            state="readonly",
            width=12
        )
        lang_combo.grid(row=0, column=1, padx=(0, 10))
        
        # 识别场景
        self.general_scene = StringVar(value="general")
        ttk.Radiobutton(param_frame, text="通用", variable=self.general_scene, value="general").grid(row=0, column=2, padx=2)
        ttk.Radiobutton(param_frame, text="印刷", variable=self.general_scene, value="print").grid(row=0, column=3, padx=2)
        ttk.Radiobutton(param_frame, text="手写", variable=self.general_scene, value="handwriting").grid(row=0, column=4, padx=2)

    def setup_table_tab(self):
        """表格识别选项卡"""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="表格识别V3")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
        
        # ===== 文件选择 =====
        file_frame = ttk.LabelFrame(tab, text="文件选择", padding="5")
        file_frame.grid(row=0, column=0, sticky=(W, E), pady=(0, 5))
        file_frame.columnconfigure(1, weight=1)
        
        # 表格文件变量
        self.table_files = []
        
        ttk.Button(
            file_frame,
            text="📁 选择图片/PDF",
            command=self.select_table_files,
            width=15
        ).grid(row=0, column=0, padx=(0, 10))
        
        self.table_file_label = ttk.Label(file_frame, text="未选择文件")
        self.table_file_label.grid(row=0, column=1, sticky=W)
        
        ttk.Button(
            file_frame,
            text="清空",
            command=self.clear_table_files,
            width=8
        ).grid(row=0, column=2, padx=(10, 0))
        
        # 文件列表
        self.table_listbox = Listbox(
            file_frame,
            height=9,
            selectmode=EXTENDED,
            activestyle='none'
        )
        self.table_listbox.grid(row=1, column=0, columnspan=3, sticky=(W, E), pady=(10, 0))

    def show_current_output(self):
        """显示当前输出目录（此方法保留但不使用）"""
        pass

    def setup_drawing_tab(self):
        """图纸图号识别选项卡"""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="图纸图号识别")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
        
        # ===== 文件选择 =====
        file_frame = ttk.LabelFrame(tab, text="PDF图纸文件", padding="5")
        file_frame.grid(row=0, column=0, sticky=(W, E), pady=(0, 5))
        file_frame.columnconfigure(1, weight=1)
        
        # 图纸文件变量
        self.drawing_files = []
        
        ttk.Button(
            file_frame,
            text="📁 选择PDF图纸",
            command=self.select_drawing_files,
            width=15
        ).grid(row=0, column=0, padx=(0, 10))
        
        self.drawing_file_label = ttk.Label(file_frame, text="未选择文件")
        self.drawing_file_label.grid(row=0, column=1, sticky=W)
        
        ttk.Button(
            file_frame,
            text="清空",
            command=self.clear_drawing_files,
            width=8
        ).grid(row=0, column=2, padx=(10, 0))
        
        # 文件列表
        self.drawing_listbox = Listbox(
            file_frame,
            height=9,
            selectmode=EXTENDED,
            activestyle='none'
        )
        self.drawing_listbox.grid(row=1, column=0, columnspan=3, sticky=(W, E), pady=(10, 0))
    
    def setup_invoice_tab(self):
        """增值税发票识别选项卡"""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="增值税发票识别")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)
        
        # ===== 文件选择 =====
        file_frame = ttk.LabelFrame(tab, text="发票图片文件", padding="5")
        file_frame.grid(row=0, column=0, sticky=(W, E), pady=(0, 5))
        file_frame.columnconfigure(1, weight=1)
        
        # 发票文件变量
        self.invoice_files = []
        
        ttk.Button(
            file_frame,
            text="📁 选择发票图片",
            command=self.select_invoice_files,
            width=15
        ).grid(row=0, column=0, padx=(0, 10))
        
        self.invoice_file_label = ttk.Label(file_frame, text="未选择文件")
        self.invoice_file_label.grid(row=0, column=1, sticky=W)
        
        ttk.Button(
            file_frame,
            text="清空",
            command=self.clear_invoice_files,
            width=8
        ).grid(row=0, column=2, padx=(10, 0))
        
        # 文件列表
        self.invoice_listbox = Listbox(
            file_frame,
            height=6,
            selectmode=EXTENDED,
            activestyle='none'
        )
        self.invoice_listbox.grid(row=1, column=0, columnspan=3, sticky=(W, E), pady=(10, 0))
        
        # ===== 输出格式设置 =====
        format_frame = ttk.LabelFrame(tab, text="输出格式设置", padding="5")
        format_frame.grid(row=1, column=0, sticky=(W, E), pady=(0, 5))
        
        ttk.Label(format_frame, text="保存格式:").grid(row=0, column=0, padx=(0, 10))
        
        ttk.Radiobutton(
            format_frame,
            text="JSON",
            variable=self.invoice_format,
            value="json"
        ).grid(row=0, column=1, padx=5)
        
        ttk.Radiobutton(
            format_frame,
            text="TXT",
            variable=self.invoice_format,
            value="txt"
        ).grid(row=0, column=2, padx=5)
        
        ttk.Radiobutton(
            format_frame,
            text="两者都保存",
            variable=self.invoice_format,
            value="both"
        ).grid(row=0, column=3, padx=5)
    
    # ========== 本地识别选项卡 ==========
    def setup_local_tab(self):
        """本地识别选项卡"""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="本地识别")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)  # 保持和原选项卡一致的行索引
        
        # ===== 服务地址设置 =====
        url_frame = ttk.LabelFrame(tab, text="服务地址", padding="5")
        url_frame.grid(row=0, column=0, sticky=(W, E), pady=(0, 10))
        url_frame.columnconfigure(1, weight=1)
        
        ttk.Label(url_frame, text="API地址:").grid(row=0, column=0, padx=(0, 5))
        ttk.Entry(url_frame, textvariable=self.local_api_url, width=50).grid(row=0, column=1, sticky=(W, E))
        ttk.Label(url_frame, text="默认: http://127.0.0.1:10001/ocr", foreground="gray").grid(row=1, column=1, sticky=W, pady=(2, 0))
        
        # ===== 文件选择 =====
        file_frame = ttk.LabelFrame(tab, text="图片文件", padding="5")
        file_frame.grid(row=1, column=0, sticky=(W, E), pady=(0, 5))
        file_frame.columnconfigure(1, weight=1)
        
        # 本地识别文件变量
        self.local_files = []
        
        ttk.Button(
            file_frame,
            text="📁 选择图片",
            command=self.select_local_files,
            width=15
        ).grid(row=0, column=0, padx=(0, 10))
        
        self.local_file_label = ttk.Label(file_frame, text="未选择文件")
        self.local_file_label.grid(row=0, column=1, sticky=W)
        
        ttk.Button(
            file_frame,
            text="清空",
            command=self.clear_local_files,
            width=8
        ).grid(row=0, column=2, padx=(10, 0))
        
        # 文件列表
        self.local_listbox = Listbox(
            file_frame,
            height=3,
            selectmode=EXTENDED,
            activestyle='none'
        )
        self.local_listbox.grid(row=1, column=0, columnspan=3, sticky=(W, E), pady=(10, 0))
        
        # ===== 连接测试按钮 =====
        test_frame = ttk.Frame(tab)
        test_frame.grid(row=2, column=0, sticky=(W, E), pady=(5, 10))
        
        ttk.Button(
            test_frame,
            text="测试连接",
            command=self.test_local_connection,
            width=15
        ).grid(row=0, column=0, padx=(0, 10))
        
        self.connection_status = ttk.Label(test_frame, text="未测试", foreground="gray")
        self.connection_status.grid(row=0, column=1, sticky=W)
    
    def select_local_files(self):
        """选择本地识别图片文件"""
        files = filedialog.askopenfilenames(
            title="选择图片文件",
            filetypes=[
                ("图像文件", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"),
                ("所有文件", "*.*")
            ]
        )
        
        if files:
            for file in files:
                if file not in self.local_files:
                    self.local_files.append(file)
                    self.local_listbox.insert(END, os.path.basename(file))
            
            self.local_file_label.config(text=f"已选择 {len(self.local_files)} 个文件")
            self.log(f"📎 本地识别文件: 已添加 {len(files)} 个文件，当前共 {len(self.local_files)} 个文件")
    
    def clear_local_files(self):
        """清空本地识别文件列表"""
        self.local_files.clear()
        self.local_listbox.delete(0, END)
        self.local_file_label.config(text="未选择文件")
        self.log("🗑️ 已清空本地识别文件列表")
    
    def test_local_connection(self):
        """测试本地服务连接"""
        try:
            api_url = self.local_api_url.get().strip()
            if not api_url:
                messagebox.showwarning("提示", "请输入服务地址")
                return
            
            self.connection_status.config(text="正在测试...", foreground="orange")
            self.root.update_idletasks()
            
            # 发送测试请求
            test_files = {'image': ('test.txt', b'test', 'text/plain')}
            response = requests.post(api_url, files=test_files, timeout=5)
            
            if response.status_code == 200:
                self.connection_status.config(text="✅ 连接成功", foreground="green")
                self.log(f"✅ 本地服务连接成功: {api_url}")
            else:
                self.connection_status.config(text=f"❌ 服务返回错误: {response.status_code}", foreground="red")
                self.log(f"❌ 本地服务返回错误: {response.status_code}")
                
        except requests.exceptions.ConnectionError:
            self.connection_status.config(text="❌ 连接失败（服务未启动）", foreground="red")
            self.log(f"❌ 无法连接到本地服务: {api_url}")
        except Exception as e:
            self.connection_status.config(text=f"❌ 测试失败", foreground="red")
            self.log(f"❌ 连接测试失败: {str(e)}")
    
    # ========== 公用方法 ==========
    def log(self, message):
        """公用日志方法"""
        self.common_log.insert(END, f"{message}\n")
        self.common_log.see(END)
        self.root.update_idletasks()
    
    def clear_log(self):
        """清除日志"""
        self.common_log.delete(1.0, END)
        self.log("📋 日志已清除")
    
    def select_common_output(self):
        global OUTPUT_DIR
        """选择公用输出目录"""
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            OUTPUT_DIR = directory
            self.log(f"📂 公用输出目录已设置为: {directory}")
    
    def clear_common_output(self):
        global OUTPUT_DIR
        """清除公用输出目录"""
        OUTPUT_DIR = "识别结果"
        self.log("🗑️ 已清除输出目录设置，将使用默认目录")
    
    # ========== 表格识别方法 ==========
    def select_table_files(self):
        """选择表格文件"""
        files = filedialog.askopenfilenames(
            title="选择图片或PDF文件",
            filetypes=[
                ("图像文件", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"),
                ("PDF文件", "*.pdf"),
                ("所有支持格式", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff *.pdf"),
                ("所有文件", "*.*")
            ]
        )
        
        if files:
            for file in files:
                if file not in self.table_files:
                    self.table_files.append(file)
                    self.table_listbox.insert(END, os.path.basename(file))
            
            self.table_file_label.config(text=f"已选择 {len(self.table_files)} 个文件")
            self.log(f"📎 表格文件: 已添加 {len(files)} 个文件，当前共 {len(self.table_files)} 个文件")
    
    def clear_table_files(self):
        """清空表格文件列表"""
        self.table_files.clear()
        self.table_listbox.delete(0, END)
        self.table_file_label.config(text="未选择文件")
        self.log("🗑️ 已清空表格文件列表")
    
    # ========== 图纸识别方法 ==========
    def select_drawing_files(self):
        """选择图纸PDF文件"""
        files = filedialog.askopenfilenames(
            title="选择PDF图纸文件",
            filetypes=[
                ("PDF文件", "*.pdf"),
                ("所有文件", "*.*")
            ]
        )
        
        if files:
            for file in files:
                if file not in self.drawing_files:
                    self.drawing_files.append(file)
                    self.drawing_listbox.insert(END, os.path.basename(file))
            
            self.drawing_file_label.config(text=f"已选择 {len(self.drawing_files)} 个文件")
            self.log(f"📎 图纸文件: 已添加 {len(files)} 个PDF图纸，当前共 {len(self.drawing_files)} 个文件")
    
    def clear_drawing_files(self):
        """清空图纸文件列表"""
        self.drawing_files.clear()
        self.drawing_listbox.delete(0, END)
        self.drawing_file_label.config(text="未选择文件")
        self.log("🗑️ 已清空图纸文件列表")
    
    # ========== 发票识别方法 ==========
    def select_invoice_files(self):
        """选择发票图片文件"""
        files = filedialog.askopenfilenames(
            title="选择发票图片文件",
            filetypes=[
                ("图像文件", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"),
                ("所有文件", "*.*")
            ]
        )
        
        if files:
            for file in files:
                if file not in self.invoice_files:
                    self.invoice_files.append(file)
                    self.invoice_listbox.insert(END, os.path.basename(file))
            
            self.invoice_file_label.config(text=f"已选择 {len(self.invoice_files)} 个文件")
            self.log(f"📎 发票文件: 已添加 {len(files)} 个文件，当前共 {len(self.invoice_files)} 个文件")
    
    def clear_invoice_files(self):
        """清空发票文件列表"""
        self.invoice_files.clear()
        self.invoice_listbox.delete(0, END)
        self.invoice_file_label.config(text="未选择文件")
        self.log("🗑️ 已清空发票文件列表")
    
    # ========== 启动识别 ==========
    def start_recognition(self):
        """根据当前选项卡启动对应的识别功能"""
        current_tab = self.notebook.index(self.notebook.select())
        
        if current_tab == 0:  # 文字识别选项卡
            self.start_general_recognition()
        elif current_tab == 1:  # 图纸识别选项卡
            self.start_drawing_recognition()
        elif current_tab == 2:  # 表格识别选项卡
            self.start_table_recognition()
        elif current_tab == 3:  # 发票识别选项卡
            self.start_invoice_recognition()
        else:  # 本地识别选项卡
            self.start_local_recognition()
    
    def start_general_recognition(self):
        """开始文字识别"""
        if not self.general_files:
            messagebox.showwarning("提示", "请选择要识别的图片文件")
            return
        
        # 禁用按钮
        self.start_btn.config(state=DISABLED)
        
        # 在新线程中执行
        thread = threading.Thread(target=self.process_general_files, daemon=True)
        thread.start()

    def process_general_files(self):
        """处理文字识别文件"""
        try:
            # 获取输出目录
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            
            # 初始化OCR识别器
            recognizer = GeneralOCRRecognizer(
                self.secret_id.get(),
                self.secret_key.get(),
                self.general_region.get()
            )
            
            total = len(self.general_files)
            self.log(f"\n{'='*50}")
            self.log(f"开始文字识别，共 {total} 个文件")
            self.log(f"输出目录: {OUTPUT_DIR}")
            self.log(f"语言类型: {self.general_language.get()}")
            self.log(f"识别场景: {self.general_scene.get()}")
            self.log(f"保存格式: {self.general_format.get()}")
            self.log(f"{'='*50}")
            
            self.common_progress['maximum'] = total
            self.common_progress['value'] = 0
            
            success = 0
            fail = 0
            
            for i, image_path in enumerate(self.general_files):
                self.log(f"\n📌 进度: {i+1}/{total}")
                
                def log_callback(msg):
                    self.log(msg)
                    self.root.update_idletasks()
                
                try:
                    result = recognizer.process_image(
                        image_path,
                        OUTPUT_DIR,
                        self.general_format.get(),
                        self.general_language.get(),
                        self.general_scene.get(),
                        log_callback
                    )
                    if result:
                        success += 1
                    else:
                        fail += 1
                except Exception as e:
                    self.log(f"❌ 处理异常: {str(e)}")
                    fail += 1
                
                self.common_progress['value'] = i + 1
                self.root.update_idletasks()
            
            self.log(f"\n{'='*50}")
            self.log(f"批量处理完成！")
            self.log(f"✅ 成功: {success} 个")
            self.log(f"❌ 失败: {fail} 个")
            self.log(f"📁 保存目录: {OUTPUT_DIR}")
            
            if success > 0:
                self.root.after(100, lambda: messagebox.showinfo(
                    "完成", 
                    f"文字识别完成！\n✅ 成功: {success} 个\n❌ 失败: {fail} 个\n\n保存位置:\n{OUTPUT_DIR}"
                ))
            
        except Exception as e:
            self.log(f"❌ 程序错误: {str(e)}")
        finally:
            self.start_btn.config(state=NORMAL)
            self.common_progress['value'] = 0

    def start_table_recognition(self):
        """开始表格识别"""
        if not self.table_files:
            messagebox.showwarning("提示", "请选择要识别的文件")
            return
        
        # 禁用按钮
        self.start_btn.config(state=DISABLED)
        
        # 在新线程中执行
        thread = threading.Thread(target=self.process_table_files, daemon=True)
        thread.start()
    
    def process_table_files(self):
        """处理表格文件"""
        try:
            recognizer = TableOCRRecognizer(
                self.secret_id.get(),
                self.secret_key.get(),
                self.table_region.get()
            )
            
            # 获取输出目录
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            
            total = len(self.table_files)
            success = 0
            fail = 0
            
            self.log(f"\n{'='*50}")
            self.log(f"开始表格识别，共 {total} 个文件")
            self.log(f"输出目录: {OUTPUT_DIR}")
            self.log(f"{'='*50}")
            
            self.common_progress['maximum'] = total
            self.common_progress['value'] = 0
            
            for i, file_path in enumerate(self.table_files):
                file_name = os.path.basename(file_path)
                self.log(f"\n[{i+1}/{total}] 处理: {file_name}")
                
                try:
                    output_path = os.path.join(
                        OUTPUT_DIR,
                        f"{Path(file_name).stem}_识别结果.xlsx"
                    )
                    
                    saved_path = recognizer.save_as_excel(file_path, output_path)
                    self.log(f"✅ 成功: {os.path.basename(saved_path)}")
                    success += 1
                    
                except Exception as e:
                    self.log(f"❌ 失败: {str(e)}")
                    fail += 1
                
                self.common_progress['value'] = i + 1
                self.root.update_idletasks()
            
            self.log(f"\n{'='*50}")
            self.log(f"处理完成！成功: {success} 个，失败: {fail} 个")
            
            if success > 0:
                self.root.after(100, lambda: messagebox.showinfo(
                    "完成", 
                    f"表格识别完成！\n成功: {success} 个\n失败: {fail} 个\n保存位置: {OUTPUT_DIR}"
                ))
            
        except Exception as e:
            self.log(f"❌ 程序错误: {str(e)}")
        finally:
            self.start_btn.config(state=NORMAL)
            self.common_progress['value'] = 0
    
    def start_drawing_recognition(self):
        """开始图纸图号识别"""
        if not self.drawing_files:
            messagebox.showwarning("提示", "请选择PDF图纸文件")
            return
        
        # 禁用按钮
        self.start_btn.config(state=DISABLED)
        
        # 在新线程中执行
        thread = threading.Thread(target=self.process_drawing_files, daemon=True)
        thread.start()
    
    def process_drawing_files(self):
        """处理图纸文件"""
        try:
            # 获取输出目录
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            setup_temp_dir()
            recognizer = DrawingNumberRecognizer(
                self.secret_id.get(),
                self.secret_key.get(),
                self.drawing_region.get()
            )
            recognizer.output_dir = OUTPUT_DIR
            
            total = len(self.drawing_files)
            self.log(f"\n{'='*50}")
            self.log(f"开始图纸图号识别，共 {total} 个文件")
            self.log(f"输出目录: {OUTPUT_DIR}")
            self.log(f"{'='*50}")
            
            self.common_progress['maximum'] = total
            self.common_progress['value'] = 0
            
            success = 0
            fail = 0
            
            for i, pdf_path in enumerate(self.drawing_files):
                self.log(f"\n📌 进度: {i+1}/{total}")
                
                def log_callback(msg):
                    self.log(msg)
                    self.root.update_idletasks()
                
                try:
                    result = recognizer.process_pdf_drawing(pdf_path, log_callback)
                    if result:
                        success += 1
                    else:
                        fail += 1
                except Exception as e:
                    self.log(f"❌ 处理异常: {str(e)}")
                    fail += 1
                
                self.common_progress['value'] = i + 1
                self.root.update_idletasks()
            
            self.log(f"\n{'='*50}")
            self.log(f"批量处理完成！")
            self.log(f"✅ 成功: {success} 个")
            self.log(f"❌ 失败: {fail} 个")
            self.log(f"📁 保存目录: {OUTPUT_DIR}")
            cleanup_temp()
            if success > 0:
                self.root.after(100, lambda: messagebox.showinfo(
                    "完成", 
                    f"图纸识别完成！\n✅ 成功命名: {success} 个\n❌ 识别失败: {fail} 个\n\n保存位置:\n{OUTPUT_DIR}"
                ))
            
        except Exception as e:
            self.log(f"❌ 程序错误: {str(e)}")
        finally:
            self.start_btn.config(state=NORMAL)
            self.common_progress['value'] = 0
    
    def start_invoice_recognition(self):
        """开始发票识别"""
        if not self.invoice_files:
            messagebox.showwarning("提示", "请选择发票图片文件")
            return
        
        # 禁用按钮
        self.start_btn.config(state=DISABLED)
        
        # 在新线程中执行
        thread = threading.Thread(target=self.process_invoice_files, daemon=True)
        thread.start()
    
    def process_invoice_files(self):
        """处理发票文件"""
        try:
            
            recognizer = VatInvoiceRecognizer(
                self.secret_id.get(),
                self.secret_key.get(),
                self.invoice_region.get()
            )
            recognizer.output_dir = OUTPUT_DIR
            
            total = len(self.invoice_files)
            self.log(f"\n{'='*50}")
            self.log(f"开始增值税发票识别，共 {total} 个文件")
            self.log(f"输出目录: {OUTPUT_DIR}")
            self.log(f"保存格式: {self.invoice_format.get()}")
            self.log(f"{'='*50}")
            
            self.common_progress['maximum'] = total
            self.common_progress['value'] = 0
            
            success = 0
            fail = 0
            
            for i, image_path in enumerate(self.invoice_files):
                self.log(f"\n📌 进度: {i+1}/{total}")
                
                def log_callback(msg):
                    self.log(msg)
                    self.root.update_idletasks()
                
                try:
                    result = recognizer.process_invoice(
                        image_path, 
                        self.invoice_format.get(),
                        log_callback
                    )
                    if result:
                        success += 1
                    else:
                        fail += 1
                except Exception as e:
                    self.log(f"❌ 处理异常: {str(e)}")
                    fail += 1
                
                self.common_progress['value'] = i + 1
                self.root.update_idletasks()
            
            self.log(f"\n{'='*50}")
            self.log(f"批量处理完成！")
            self.log(f"✅ 成功: {success} 个")
            self.log(f"❌ 失败: {fail} 个")
            self.log(f"📁 保存目录: {OUTPUT_DIR}")
            
            if success > 0:
                self.root.after(100, lambda: messagebox.showinfo(
                    "完成", 
                    f"发票识别完成！\n✅ 成功: {success} 个\n❌ 失败: {fail} 个\n\n保存位置:\n{OUTPUT_DIR}"
                ))
            
        except Exception as e:
            self.log(f"❌ 程序错误: {str(e)}")
        finally:
            self.start_btn.config(state=NORMAL)
            self.common_progress['value'] = 0
    
    # ========== 本地识别方法 ==========
    def start_local_recognition(self):
        """开始本地识别"""
        if not self.local_files:
            messagebox.showwarning("提示", "请选择要识别的图片文件")
            return
        
        api_url = self.local_api_url.get().strip()
        if not api_url:
            messagebox.showwarning("提示", "请输入本地服务地址")
            return
        
        # 禁用按钮
        self.start_btn.config(state=DISABLED)
        
        # 在新线程中执行
        thread = threading.Thread(target=self.process_local_files, daemon=True)
        thread.start()
    
    def process_local_files(self):
        """处理本地识别文件"""
        try:
            # 获取输出目录
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            
            # 初始化本地识别器
            recognizer = LocalOCRRecognizer(self.local_api_url.get().strip())
            
            total = len(self.local_files)
            self.log(f"\n{'='*50}")
            self.log(f"开始本地识别，共 {total} 个文件")
            self.log(f"输出目录: {OUTPUT_DIR}")
            self.log(f"服务地址: {recognizer.api_url}")
            self.log(f"{'='*50}")
            
            self.common_progress['maximum'] = total
            self.common_progress['value'] = 0
            
            success = 0
            fail = 0
            
            for i, image_path in enumerate(self.local_files):
                self.log(f"\n📌 进度: {i+1}/{total}")
                
                def log_callback(msg):
                    self.log(msg)
                    self.root.update_idletasks()
                
                try:
                    result = recognizer.process_image(
                        image_path,
                        OUTPUT_DIR,
                        log_callback
                    )
                    if result:
                        success += 1
                    else:
                        fail += 1
                except Exception as e:
                    self.log(f"❌ 处理异常: {str(e)}")
                    fail += 1
                
                self.common_progress['value'] = i + 1
                self.root.update_idletasks()
            
            self.log(f"\n{'='*50}")
            self.log(f"批量处理完成！")
            self.log(f"✅ 成功: {success} 个")
            self.log(f"❌ 失败: {fail} 个")
            self.log(f"📁 保存目录: {OUTPUT_DIR}")
            
            if success > 0:
                self.root.after(100, lambda: messagebox.showinfo(
                    "完成", 
                    f"本地识别完成！\n✅ 成功: {success} 个\n❌ 失败: {fail} 个\n\n保存位置:\n{OUTPUT_DIR}"
                ))
            
        except Exception as e:
            self.log(f"❌ 程序错误: {str(e)}")
        finally:
            self.start_btn.config(state=NORMAL)
            self.common_progress['value'] = 0


# ==================== 程序入口 ====================
def main():
    """主函数"""
    root = Tk()
    app = OCRTabbedApp(root)
    
    # 窗口居中
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()


if __name__ == "__main__":
    main()